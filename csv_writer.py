# -*- coding: utf-8 -*-
"""

This module is used to write information into excel

Attributes
----------
NEW_ISS_TEMP_FIXED_FIELDS : dict
    fixed-value fields on new_issuance_template
NEW_ISS_TEMP_FIELDS_MAPPING : dict
    mapping of fields on new_issuance_template to fields on sail template

"""

import pandas as pd

from datetime import datetime
from openpyxl import load_workbook
from preprocess import to_index, convert_waterfallfreq

NEW_ISS_TEMP_FIXED_FIELDS = {"COLLATERAL GROUP": "all",
                             "CURRENCY": "CNY",
                             "ACCRUE ON SHORTFALL?": "N",
                             "IMPLIED LOSS?": "N",
                             "PAYMENT DELAY": "0",
                             "DATE ROLL": "Following",
                             "DAY COUNT": "ACTUAL_365",
                             "CALENDAR": "CH",
                             "RECORD DELAY": "0",
                             "PRICE AT ISSUANCE": "100",
                             "CSDC?": "Y",
                             "EXCHANGE STATUS": "confirmed"}
NEW_ISS_TEMP_FIELDS_MAPPING = {"CLASS NAME": "ClsName",
                               "ORIGINAL BALANCE": "TrancheOriBal",
                               "DATED DATE MM/DD/YYYY": "SetDt",
                               "SETTLE DATE MM/DD/YYYY": "SetDt",
                               "PRICING DATE MM/DD/YYYY": "SetDt",
                               "FIRST PAYMENT DATE MM/DD/YYYY": "FirPayDt",
                               "MATURITY DATE MM/DD/YYYY": "FinalMty",
                               "ID: CHINA ID ": "ChinaID",
                               "ORIGINAL COUPON": "Cpn",
                               "EXCHANGE STATUS DATE MM/DD/YYYY": "LstDt",
                               "ACCRUAL#1 COUPON": "Cpn",
                               "EXPECTED MATURITY DATE MM/DD/YYYY": "ExpMty",
                               "INTEREST FREQUENCY": "IntFreq",
                               "PRINCIPAL FREQUENCY": "IntFreq",
                               "EXCHANGE NAME": "Exch",
                               "CLASS DESCRIPTION": "ClsDes"
                               }


class Writer(object):
    def __init__(self, input_filepath, output_filepath, deal):

        self.input_filepath = input_filepath
        self.output_filepath = output_filepath
        self.workbook = load_workbook(self.input_filepath, data_only=True)
        self.deal = deal
        self.tranches = deal["tranches"] if deal.has_key("tranches") else None

    def __load(self, sheet):

        return self.workbook[sheet]

    def __save(self):

        self.workbook.save(self.output_filepath)

    def write2clsinfo(self, sheet):
        """write to classinfo sheet"""
        worksheet = self.__load(sheet)
        fields_name = [cell.value for cell in worksheet['A']]
        index_mapping = to_index(NEW_ISS_TEMP_FIELDS_MAPPING, fields_name, "list")
        if self.tranches:
            for index, tranche in enumerate(self.tranches):
                for key in index_mapping.keys():
                    if tranche.has_key(key):
                        for i in xrange(len(index_mapping[key])):
                            worksheet.cell(index_mapping[key][i] + 1, 2 + index).value = tranche[key]
                    elif self.deal.has_key(key):
                        for i in xrange(len(index_mapping[key])):
                            worksheet.cell(index_mapping[key][i] + 1, 2 + index).value = self.deal[key]
                for key in NEW_ISS_TEMP_FIXED_FIELDS.keys():
                    worksheet.cell(fields_name.index(key) + 1, 2 + index).value = NEW_ISS_TEMP_FIXED_FIELDS[key]
                # csort
                worksheet.cell(5, 2 + index).value = index + 1
                # supporting class
                worksheet.cell(16, 2 + index).value = self.tranches[index + 1]["ClsName"] if index + 1 <= len(
                    self.tranches) - 1 else 0
        self.__save()

    def write2collainfo(self, sheet):
        """write to collateralinfo sheet"""
        worksheet = self.__load(sheet)
        if self.deal.has_key("DealAmount"):
            worksheet.cell(5, 2).value = self.deal["DealAmount"]
            worksheet.cell(5, 3).value = self.deal["DealAmount"]
        self.__save()

    def write2dealinfo(self, sheet):
        """write to dealinfo sheet"""
        worksheet = self.__load(sheet)
        worksheet.cell(8, 2).value = "abs"
        if self.tranches:
            worksheet.cell(11, 2).value = self.tranches[0]["SetDt"]
            setdt = pd.to_datetime(self.tranches[0]["SetDt"])
            worksheet.cell(12, 2).value = datetime(setdt.year, setdt.month, 1, 00, 00, 00)
            worksheet.cell(13, 2).value = convert_waterfallfreq(self.tranches[0]["IntFreq"])
        self.__save()

    def write2relatedPar(self, sheet):
        """write to relateParty sheet"""
        worksheet = self.__load(sheet)
        if self.deal.has_key("LM"):
            for index, LM in enumerate(self.deal['LM']):
                worksheet.cell(4, 2 + index).value = "DEAL"
                worksheet.cell(5, 2 + index).value = LM
                worksheet.cell(6, 2 + index).value = 1.0 / len(self.deal['LM'])
        if self.deal.has_key("AssetM"):
            for index, AssetM in enumerate(self.deal['AssetM']):
                worksheet.cell(46, 2 + index).value = "DEAL"
                worksheet.cell(47, 2 + index).value = AssetM
                worksheet.cell(48, 2 + index).value = 1.0 / len(self.deal['AssetM'])
        self.__save()
