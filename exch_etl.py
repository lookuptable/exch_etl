# -*- coding: utf-8 -*-
"""
Created on Thu Aug 01 08:42:28 2019

@author: sliu439
"""


import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import numpy as np

from entity_mapping import entity_mapping
from msg import msg
import bcos


SAIL_FIELDS_MAPPING={ "Deal Full Name":"DealName",
                     "DealID":"DealID",
                    "EXCHANGE NAME":"Exch",
                    "Type":"DealType",
                    "Asset Manager": "AssetM",
                    "Lead Manager":"LM",
                    "Originator":"Originator",
                    "Class Name":"ClsName",
                    "Chinese ID":"ChineseID",
                    "CHINA ID ":"ChinaID",
                    "Original Balance":"TrancheOriBal",
                    "Issue Amount":"DealAmount",
                    "SETTLEMENT DATE":"SetDt",
                    "FIRST PAYMENT DATE":"FirPayDt",
                    "MATURITY DATE":"FinalMty",
                    "Expected MATURITY DATE":"ExpMty",
                    "ORIGINAL COUPON":"Cpn",
                    "INTEREST FREQUENCY":"IntFreq",
                    "PRINCIPAL FREQUENCY":"PrinFreq",
                    "LISTING DATE":"LstDt",
                    "Rating":"Rating",
                    "Rating Agency":"RtAgency"
                    }
TRANCH_COLS =["ClsName","ChineseID","ChinaID",
             "TrancheOriBal","SetDt",
             "FirPayDt","ExpMty","FinalMty",
             "Cpn","IntFreq","ClsDes"
             ]
             
NEW_ISS_TEMP_FIELDS_MAPPING={"CLASS NAME":"ClsName",
                             "ORIGINAL BALANCE":"TrancheOriBal",
                             "DATED DATE MM/DD/YYYY":"SetDt",
                             "SETTLE DATE MM/DD/YYYY":"SetDt",
                             "PRICING DATE MM/DD/YYYY":"SetDt",
                             "FIRST PAYMENT DATE MM/DD/YYYY":"FirPayDt",
                             "MATURITY DATE MM/DD/YYYY":"FinalMty",
                             "ID: CHINA ID ":"ChinaID",
                             "ORIGINAL COUPON":"Cpn",
                             "EXCHANGE STATUS DATE MM/DD/YYYY":"LstDt",
                             "ACCRUAL#1 COUPON":"Cpn",
                             "EXPECTED MATURITY DATE MM/DD/YYYY":"ExpMty",
                             "INTEREST FREQUENCY":"IntFreq",
                             "PRINCIPAL FREQUENCY":"IntFreq",
                             "EXCHANGE NAME":"Exch",
                             "CLASS DESCRIPTION":"ClsDes"
                             }
NEW_ISS_TEMP_FIXED_FIELDS={"COLLATERAL GROUP":"all",
                           "CURRENCY":"CNY",
                           "ACCRUE ON SHORTFALL?":"N",
                           "IMPLIED LOSS?":"N",
                           "PAYMENT DELAY":"0",
                           "DATE ROLL":"Following",
                           "DAY COUNT":"ACTUAL_365",
                           "CALENDAR":"CH",
                           "RECORD DELAY":"0",
                           "PRICE AT ISSUANCE":"100",
                           "CSDC?":"Y",
                           "EXCHANGE STATUS":"confirmed"}      

    
def load_sail_template(filepath):
    
    if filepath.split(".")[-1] =="xlsx":
        return pd.read_excel(filepath)
    else:
        print "file format wrong"

def to_index(mapping,headers,rtype):
    
    index_mapping={}
    if rtype == "str":
        for key in mapping.keys():
            if key in headers:
                    index_mapping[mapping[key]]=headers.index(key)
    elif rtype == "list":
        for key in mapping.keys():
            if key in headers:
                    if mapping[key] in index_mapping.keys():
                        index_mapping[mapping[key]].append(headers.index(key))
                    else: 
                        index_mapping[mapping[key]]=[headers.index(key)]      
    return index_mapping
    
def groupbydataframe(df,column="DealID"):
    
    #for key, value in df.groupby(column):
    #    yield value.reset_index(drop=True)
    return (value.reset_index(drop=True) for key, value in df.groupby(column))
    
def convertFreq(string):
    
    res=""
    if string == u"季付":
        res="Q"    
    elif string == u"半年付":
        res="SA"  
    elif string == u"年付" or u"到期偿还":
        res="A"
    elif string == u"月付":
        res="M"
    return res
    
def convertWaterfallFreq(string):
    
    res=""
    if string == "Q":
        res= 4
    elif string == "A":
        res=1
    elif string == "M":
        res=12
    elif string == "SA":
        res=2 
    return res
    
def returnClsDes(FirPayDt,FinalMty,ClsName):
    
    res=""
    if ClsName.upper()=="SUB":
        res="SUB"
    elif FirPayDt==FinalMty:
        res="HB"
    else:
        res="PT,SEQ"
    return res
    
def convertExch(string):
    
    res=""
    if string == u"上海证券交易所":
        res="Shanghai"
    elif string == u"深圳证券交易所":
        res="Shenzhen"
    elif string == u"银行间债券市场":
        res="China Interbank" 
    return res

def mapEntity(string):
    
    return finder.get_eng_name(string)
    
def normalize(df,index_mapping):
    
    df=df.fillna("")
    deal={}
    tranches=[]
    for index,row in df.iterrows():
            tranche={}
            if index ==0:
                for key in index_mapping.keys():
                    if key not in TRANCH_COLS:
                       if key == "Exch":
                           deal[key]=convertExch(row.iloc[index_mapping[key]])
                       elif key == "AssetM" or key == "LM":
                           deal[key]=[mapEntity(ele) for ele in row.iloc[index_mapping[key]].split(r",")] 
                       elif key in ["LstDt"]:
                           try:
                               deal[key] = datetime.strptime(row.iloc[index_mapping[key]],'%m/%d/%Y')
                           except ValueError:
                               deal[key] = ""
                               print "please check data format of {0}:{1}".format(key,row.iloc[index_mapping[key]])
                       else:
                           deal[key]=row.iloc[index_mapping[key]]
            for key in TRANCH_COLS:
                if index_mapping.has_key(key):
                   if key == "IntFreq":
                      tranche[key]=convertFreq(row.iloc[index_mapping[key]])
                   elif key in ["FirPayDt","FinalMty","LstDt","ExpMty","SetDt"]:
                      try:
                          tranche[key] = datetime.strptime(row.iloc[index_mapping[key]],'%m/%d/%Y')
                      except ValueError:
                          tranche[key] = ""
                          print "please check data format of {0}:{1}".format(key,row.iloc[index_mapping[key]])
                   else:
                       tranche[key]=row.iloc[index_mapping[key]]  
            if tranche.has_key("FirPayDt") and tranche.has_key("FinalMty") and tranche.has_key("ClsName"):
                tranche["ClsDes"]=returnClsDes(tranche["FirPayDt"],tranche["FinalMty"],tranche["ClsName"]) 
            tranches.append(tranche)
    deal["tranches"]=tranches
    return deal

class writer(object):
    
    def __init__(self,input_filepath,output_filepath,deal):
        
        self.input_filepath=input_filepath
        self.output_filepath=output_filepath
        self.workbook=load_workbook(self.input_filepath)
        self.deal = deal
        self.tranches = deal["tranches"] if deal.has_key("tranches") else None
        
    def load (self,sheet):
     
        return self.workbook[sheet]
        
    def save(self):
        
         self.workbook.save(self.output_filepath) 
         
    def write2clsinfo(self,sheet):
        
        worksheet = self.load(sheet)
        fields_name=[cell.value for cell in worksheet['A']]
        index_mapping=to_index(NEW_ISS_TEMP_FIELDS_MAPPING,fields_name,"list")
        #print index_mapping
        if self.tranches:
          for index, tranche in enumerate(self.tranches):
                for key in index_mapping.keys():
                    if tranche.has_key(key):
                        for i in xrange(len(index_mapping[key])):
                            worksheet.cell(index_mapping[key][i]+1,2+index).value=tranche[key]
                    elif self.deal.has_key(key):
                        for i in xrange(len(index_mapping[key])):
                            worksheet.cell(index_mapping[key][i]+1,2+index).value=self.deal[key]                        
                #fixed fields
                for key in NEW_ISS_TEMP_FIXED_FIELDS.keys():
                        worksheet.cell(fields_name.index(key)+1,2+index).value=NEW_ISS_TEMP_FIXED_FIELDS[key]
            #rest of fields
                #csort
                worksheet.cell(5,2+index).value=index+1   
                #supporting class
                worksheet.cell(16,2+index).value=self.tranches[index+1]["ClsName"] if index+1 <= len(self.tranches)-1 else 0
        self.save()
                     
    def write2collainfo(self,sheet):
        
        worksheet = self.load(sheet)
        if deal.has_key("DealAmount"):
            worksheet.cell(5,2).value = self.deal["DealAmount"]
            worksheet.cell(5,3).value = self.deal["DealAmount"] 
        self.save()
        
    def write2dealinfo(self,sheet):
        
        worksheet = self.load(sheet)
        worksheet.cell(8,2).value = "abs"
        if self.tranches:
            worksheet.cell(11,2).value = self.tranches[0]["SetDt"]
            setdt = pd.to_datetime(self.tranches[0]["SetDt"])
            worksheet.cell(12,2).value = datetime(setdt.year,setdt.month,1,00,00,00)
            worksheet.cell(13,2).value = convertWaterfallFreq(self.tranches[0]["IntFreq"])
        self.save()
     
    def write2relatedPar(self,sheet):
         
        worksheet = self.load(sheet)
        if deal.has_key("LM"):
            for index,LM in enumerate(self.deal['LM']):
                worksheet.cell(4,2+index).value="DEAL"
                worksheet.cell(5,2+index).value=LM
                worksheet.cell(6,2+index).value=1.0/len(self.deal['LM'])
        if deal.has_key("AssetM"): 
            for index,AssetM in enumerate(self.deal['AssetM']):
                worksheet.cell(46,2+index).value="DEAL"
                worksheet.cell(47,2+index).value=AssetM
                worksheet.cell(48,2+index).value=1.0/len(self.deal['AssetM'])
        self.save()
         

if __name__ == "__main__":
    
    finder=entity_mapping()
    msger=msg('sliu439@bloomberg.net','apmort@bloomberg.net')
    bcosinfo=bcos.BcosInfo("gd-mtge","gd-mtge-bcos-account","1153f598e3e026a8e710d384dbd7483b0d3724d42c2bd937ccb83bb7f469ff85")
    res=bcos.monitor("CN_NEWS_AUTOMATION",bcosinfo,900,1,'windows')
    for i in res:

        if bcos.scan(".*_-Template-.*.xlsx",i):
           try:
               flag,filepath=bcos.download(i,bcosinfo,r"F:\MTGE\automation\exch_etl\dailyfile\\"+str(i).split(r"/")[-1],1,'windows')
               if flag:   
                   df=load_sail_template(filepath)
                   index_mapping_sail= to_index(SAIL_FIELDS_MAPPING,df.columns.tolist(),"str")
                   for df in groupbydataframe(df):
                       deal =normalize(df,index_mapping_sail)
                       excel_writer=writer(r"F:\MTGE\automation\exch_etl\issuance_new_deal_template.xlsx",r"F:\MTGE\automation\exch_etl\dailyfile\issuance_new_deal_template_"+deal["DealName"]+".XLSX",deal)
                       excel_writer.write2clsinfo("Class Information")
                       excel_writer.write2collainfo("Collateral Information")
                       excel_writer.write2dealinfo("Deal Information")
                       excel_writer.write2relatedPar("Related Parties")
                       msger.send("SAIL Daily File - {0}".format(deal["DealName"].encode('utf-8')),'Please find attached deal',r"F:\MTGE\automation\exch_etl\dailyfile\issuance_new_deal_template_"+deal["DealName"]+".XLSX",deal["DealName"].encode('utf-8')+"_issuance_new_deal_template.XLSX")
           except Exception as e:
               print e.message
    